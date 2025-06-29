# 必须包含excel的文本内容，否则llm容易将繁体字直接转换为简体字
import base64
import re
from io import BytesIO
from pathlib import Path

from PIL import Image
from dotenv import load_dotenv, find_dotenv
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.prompts.chat import _TextTemplateParam, HumanMessagePromptTemplate, _ImageTemplateParam
from langchain_openai import ChatOpenAI

system_no_example = """
你是专业python编码助手。
请根据以下指引，修改python代码，使用openpyxl处理指定文件路径{file_path}的excel文件。

# 处理步骤
1. 逐行解析python代码，针对可能存在错误或漏洞的地方，修改代码逻辑。
2. 结合思考内容{thought}，修改代码逻辑。

# 待修改python代码
```
{python_code}
```

请始终遵循以上指引调整python代码。
"""

_ = load_dotenv(find_dotenv())

def image_to_base64(image_path: Path) -> str:
    image = Image.open(image_path)
    buffered = BytesIO()
    image.save(buffered, format=image.format if image.format is not None else "JPEG")
    return base64.b64encode(buffered.getvalue()).decode("utf-8")


def extract_python_code(markdown_string: str) -> str:
    # Strip whitespace to avoid indentation errors in LLM-generated code
    markdown_string = markdown_string.strip()

    # Regex pattern to match Python code blocks
    pattern = r"```[\w\s]*python\n([\s\S]*?)```|```([\s\S]*?)```"

    # Find all matches in the markdown string
    matches = re.findall(pattern, markdown_string, re.IGNORECASE)

    # Extract the Python code from the matches
    python_code = []
    for match in matches:
        python = match[0] if match[0] else match[1]
        python_code.append(python.strip())

    if len(python_code) == 0:
        return markdown_string

    return python_code[0]


PROMPT_TEMPLATE = ChatPromptTemplate.from_messages([
        (
            "system",
            system_no_example,
        ),
        HumanMessagePromptTemplate.from_template([
            # _TextTemplateParam(text="这是文档当前页面的文本内容：{content}"),
            _TextTemplateParam(text="这是指定文件路径的excel文件截图："),
            _ImageTemplateParam(image_url="data:image/JPEG;base64,{image}"),
        ]),
        (
            "system",
            "{thought}",
        )
    ])

chain = (
        PROMPT_TEMPLATE
        | ChatOpenAI(model="gpt-4o", temperature=0, api_key="xxxx", base_url="xxxx")
)

response = chain.invoke({
    "file_path": Path("E:\document\CASH相关\Remittance文件\PaymentAdvice20241106170146（客户发出的）.xlsx"),
    "image": image_to_base64(Path("E:\document\CASH相关\Remittance文件\PaymentAdvice20241106170146（客户发出的）.png")),
    "thought": "执行python代码的最终结果数据中关联订单信息associated_orders为空，为什么associated_orders为空？如何修改代码逻辑，确保关联订单信息associated_orders不为空。",
    "python_code":
"""
import openpyxl

def extract_payment_data(file_path):
    # 加载Excel文件
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active

    # 初始化结果字典
    payment_info = {{
        "payment_customer_name": "",
        "payment_date": "",
        "payment_amount": "",
        "payment_reference_code": ""
    }}
    associated_orders = []

    # 遍历表格数据
    for row in sheet.iter_rows(values_only=True):
        # 提取付款信息
        if row[0] == "付款對象" and "聯想" not in str(row[1]):
            payment_info["payment_customer_name"] = str(row[1]).strip()
        elif row[0] == "付款金額":
            payment_info["payment_amount"] = str(row[1]).replace("NTD", "").strip()
        elif row[0] == "付款日":
            payment_info["payment_date"] = str(row[1]).strip()
        elif row[0] == "合計":
            payment_info["payment_reference_code"] = str(sheet.cell(row=row[0].row, column=1).value).strip()

        # 提取关联订单信息
        if isinstance(row[0], int):  # 检查是否为订单数据行
            billing_reference = str(row[2]).strip()
            paid_amount = str(row[3]).strip()
            if billing_reference.startswith("EY") or billing_reference.startswith("AZ") or billing_reference.startswith("ZA"):
                associated_orders.append({{
                    "so_code": "",
                    "billing_code": "",
                    "billing_reference": billing_reference,
                    "tds_amount": "",
                    "vat_amount": "",
                    "wht_amount": "",
                    "paid_amount": paid_amount
                }})

    # 返回提取结果
    return payment_info, associated_orders


# 文件路径
file_path = r"E:\document\CASH相关\Remittance文件\PaymentAdvice20241106170146（客户发出的）.xlsx"

# 提取数据
payment_info, associated_orders = extract_payment_data(file_path)

# 打印结果
print("付款信息:")
print(payment_info)
print("\n关联订单信息:")
for order in associated_orders:
    print(order) 
"""
})

# If no response with text is found, return the first response's content (which may be empty)
result = response.content
print(result, end="-------------------------\n")
plotly_code = extract_python_code(result)

print("Extracted Python code: \n" + plotly_code)
ldict = {
    "file_path": Path("E:\document\CASH相关\Remittance文件\PaymentAdvice20241106170146（客户发出的）.xlsx")
}
try:
    import openpyxl
    exec(plotly_code, globals(), ldict)
    print(ldict)
except Exception as e:
    raise RuntimeError(f"Error executing code: {e}")
